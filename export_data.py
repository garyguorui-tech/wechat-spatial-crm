# -*- coding: utf-8 -*-
"""把 wechat_data.json 导出成 Excel(.xlsx) 和 CSV(UTF-8 BOM)，方便直接查看。"""
import csv
import json
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.join(HERE, "wechat_data.json")
XLSX = os.path.join(HERE, "客情数据.xlsx")
CSV = os.path.join(HERE, "客情数据.csv")

# 列定义：(表头, 取值函数)
COLUMNS = [
    ("序号",       lambda i, r: i + 1),
    ("昵称",       lambda i, r: r.get("nickname", "")),
    ("备注",       lambda i, r: r.get("remark", "")),
    ("地区",       lambda i, r: r.get("region", "")),
    ("标签",       lambda i, r: r.get("tags", "")),
    ("微信号",     lambda i, r: r.get("wxid", "")),
    ("最后对话时间", lambda i, r: r.get("last_chat_time", "")),
    ("是否合作伙伴", lambda i, r: "是" if r.get("is_partner") else "否"),
    ("最近消息",   lambda i, r: " / ".join(r.get("chat_history", []))),
    ("纬度",       lambda i, r: r.get("lat", "")),
    ("经度",       lambda i, r: r.get("lng", "")),
]


def main():
    data = json.load(open(DATA, encoding="utf-8"))

    # ---------- 导出 Excel ----------
    wb = Workbook()
    ws = wb.active
    ws.title = "客情数据"

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    partner_fill = PatternFill("solid", fgColor="E2EFDA")   # 合作伙伴行浅绿底

    # 表头
    for col, (title, _) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=1, column=col, value=title)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")

    # 数据行
    for i, rec in enumerate(data):
        row = i + 2
        for col, (_, getter) in enumerate(COLUMNS, start=1):
            ws.cell(row=row, column=col, value=getter(i, rec))
        if rec.get("is_partner"):                          # 合作伙伴整行高亮
            for col in range(1, len(COLUMNS) + 1):
                ws.cell(row=row, column=col).fill = partner_fill

    # 列宽
    widths = [6, 22, 26, 16, 16, 16, 14, 12, 24, 10, 10]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"                                  # 冻结表头
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"  # 表头加筛选
    wb.save(XLSX)

    # ---------- 导出 CSV（UTF-8 BOM，Excel 打开不乱码） ----------
    with open(CSV, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow([t for t, _ in COLUMNS])
        for i, rec in enumerate(data):
            w.writerow([g(i, rec) for _, g in COLUMNS])

    partners = sum(1 for r in data if r.get("is_partner"))
    print(f"共 {len(data)} 条（合作伙伴 {partners}）")
    print(f"Excel: {XLSX}")
    print(f"CSV  : {CSV}")


if __name__ == "__main__":
    main()
